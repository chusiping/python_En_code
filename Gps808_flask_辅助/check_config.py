import re
from openpyxl import load_workbook
from collections import defaultdict

xlsx_path = r'E:\git_15home\python_En_code_git\Gps808_flask_辅助\config_20260324223310.xlsx'
output_path = r'E:\git_15home\python_En_code_git\Gps808_flask_辅助\rt.txt'

wb = load_workbook(xlsx_path)
ws = wb['Sheet1']

headers = [cell.value for cell in ws[1]]
rows = list(ws.iter_rows(min_row=2, values_only=True))

stats = defaultdict(lambda: {'total': 0, 'errors': []})
error_list = []

def validate(row_idx, row):
    r = dict(zip(headers, row))
    errors = []

    if not r['name'] or not str(r['name']).isdigit():
        errors.append(f"行{row_idx+1}: name必须为数字")

    if r['excel_file']:
        path = str(r['excel_file'])
        if not path.startswith('excle\\') or not path.endswith('.xlsx'):
            errors.append(f"行{row_idx+1}: excel_file格式错误 '{path}'")
        elif not re.match(r'^excle\\[\u4e00-\u9fa5a-zA-Z0-9_\-\s]+\.xlsx$', path):
            errors.append(f"行{row_idx+1}: excel_file包含非法字符 '{path}'")

    if r['server_ip']:
        ip = str(r['server_ip'])
        if not re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', ip):
            errors.append(f"行{row_idx+1}: server_ip格式错误 '{ip}'")
        else:
            parts = ip.split('.')
            if not all(0 <= int(p) <= 255 for p in parts):
                errors.append(f"行{row_idx+1}: server_ip超出范围 '{ip}'")

    if r['server_port']:
        port = r['server_port']
        if not isinstance(port, int) or not (1 <= port <= 65535):
            errors.append(f"行{row_idx+1}: server_port必须为1-65535整数")

    if r['terminal_phone']:
        phone = str(r['terminal_phone'])
        if not phone.isdigit() or len(phone) != 11:
            errors.append(f"行{row_idx+1}: terminal_phone必须为11位数字")

    if r['start_date']:
        date = str(r['start_date'])
        if not re.match(r'^\d{4}-\d{1,2}-\d{1,2}$', date):
            errors.append(f"行{row_idx+1}: start_date格式错误 '{date}'")

    if r['start_time']:
        time_val = str(r['start_time'])
        if not time_val.startswith('=') or 'RANDBETWEEN' not in time_val:
            errors.append(f"行{row_idx+1}: start_time应为随机时间公式")

    if r['enabled'] is not None:
        enabled = str(r['enabled'])
        if enabled not in ['0', '1', '√', '是', '启用']:
            errors.append(f"行{row_idx+1}: enabled值异常 '{enabled}'")

    return errors

for idx, row in enumerate(rows):
    row_errors = validate(idx + 2, row)
    if row_errors:
        error_list.extend(row_errors)
    for h, v in zip(headers, row):
        stats[h]['total'] += 1
        if v is not None:
            stats[h]['valid'] = stats[h].get('valid', 0) + 1

with open(output_path, 'w', encoding='utf-8') as f:
    f.write("=" * 60 + "\n")
    f.write("配置数据校验报告\n")
    f.write("=" * 60 + "\n\n")

    f.write("【数据概览】\n")
    f.write(f"总行数: {len(rows)}\n")
    f.write(f"总列数: {len(headers)}\n\n")

    f.write("【字段统计】\n")
    for h in headers:
        s = stats[h]
        valid = s.get('valid', 0)
        f.write(f"{h}: {valid}/{s['total']} 非空\n")

    f.write("\n【IP统计】\n")
    ip_count = defaultdict(int)
    for row in rows:
        if row[headers.index('server_ip')]:
            ip_count[row[headers.index('server_ip')]] += 1
    for ip, cnt in sorted(ip_count.items(), key=lambda x: -x[1]):
        f.write(f"  {ip}: {cnt}条\n")

    f.write("\n【端口统计】\n")
    port_count = defaultdict(int)
    for row in rows:
        if row[headers.index('server_port')]:
            port_count[row[headers.index('server_port')]] += 1
    for port, cnt in sorted(port_count.items(), key=lambda x: -x[1]):
        f.write(f"  {port}: {cnt}条\n")

    f.write("\n【终端电话统计】\n")
    phone_count = defaultdict(int)
    for row in rows:
        if row[headers.index('terminal_phone')]:
            phone_count[row[headers.index('terminal_phone')]] += 1
    for phone, cnt in sorted(phone_count.items(), key=lambda x: -x[1]):
        f.write(f"  {phone}: {cnt}条\n")

    f.write("\n【号码对应Excel文件统计】\n")
    phone_xlsx_count = defaultdict(set)
    for row in rows:
        phone = row[headers.index('terminal_phone')]
        xlsx = row[headers.index('excel_file')]
        if phone and xlsx:
            phone_xlsx_count[phone].add(xlsx)
    for phone, xlsx_set in sorted(phone_xlsx_count.items(), key=lambda x: -len(x[1])):
        f.write(f"  {phone}: {len(xlsx_set)}个文件\n")
        for xf in sorted(xlsx_set):
            f.write(f"    - {xf}\n")

    f.write("\n【按日期汇总统计】\n")
    date_count = defaultdict(int)
    for row in rows:
        date = row[headers.index('start_date')]
        if date:
            date_count[date] += 1
    for date, cnt in sorted(date_count.items()):
        f.write(f"  {date}: {cnt}条\n")

    f.write("\n【按日期+号码交叉统计】\n")
    date_phone_count = defaultdict(lambda: defaultdict(int))
    for row in rows:
        date = row[headers.index('start_date')]
        phone = row[headers.index('terminal_phone')]
        if date and phone:
            date_phone_count[date][phone] += 1
    for date in sorted(date_phone_count.keys()):
        f.write(f"  {date}:\n")
        for phone, cnt in sorted(date_phone_count[date].items()):
            f.write(f"    {phone}: {cnt}条\n")

    f.write("\n【启用状态统计】\n")
    enabled_count = defaultdict(int)
    for row in rows:
        e = row[headers.index('enabled')]
        enabled_count[str(e) if e is not None else '空'] += 1
    for s, cnt in sorted(enabled_count.items()):
        f.write(f"  {s}: {cnt}条\n")

    if error_list:
        f.write("\n【错误详情】\n")
        for e in error_list:
            f.write(f"  {e}\n")
        f.write(f"\n共发现 {len(error_list)} 个错误\n")
    else:
        f.write("\n【校验结果】\n")
        f.write("  未发现格式错误\n")

print(f"校验完成，结果已写入 {output_path}")