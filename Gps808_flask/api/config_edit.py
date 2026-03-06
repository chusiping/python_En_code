from flask import Blueprint, request, jsonify, session, redirect
import openpyxl
import os
from datetime import datetime, date, time

config_bp = Blueprint('config', __name__)

CONFIG_DIR = 'config'

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated_function

def get_excel_files():
    if not os.path.exists(CONFIG_DIR):
        return []
    return [f for f in os.listdir(CONFIG_DIR) if f.endswith(('.xlsx', '.xls'))]

def get_json_files():
    if not os.path.exists(CONFIG_DIR):
        return []
    return [f for f in os.listdir(CONFIG_DIR) if f.endswith('.json')]

def convert_value(val):
    if val is None:
        return ''
    if isinstance(val, (datetime, date, time)):
        return val.isoformat()
    return val

def get_workbook(filename=None):
    if filename is None:
        filename = 'config.xlsx'
    filepath = os.path.join(CONFIG_DIR, filename)
    if not os.path.exists(filepath):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet('Sheet1')
        wb.save(filepath)
    return openpyxl.load_workbook(filepath)

def save_workbook(wb, filename=None):
    if filename is None:
        filename = 'config.xlsx'
    filepath = os.path.join(CONFIG_DIR, filename)
    wb.save(filepath)

@config_bp.route('/files', methods=['GET'])
@login_required
def list_files():
    files = get_excel_files()
    return jsonify(files)

@config_bp.route('/jsonfiles', methods=['GET'])
@login_required
def list_json_files():
    files = get_json_files()
    return jsonify(files)

@config_bp.route('/data', methods=['GET'])
@login_required
def get_data():
    filename = request.args.get('file', 'config.xlsx')
    wb = get_workbook(filename)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        row_data = [convert_value(cell) for cell in row]
        if any(cell != '' for cell in row_data):
            data.append(row_data)
    wb.close()
    response = jsonify(data)
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return response

@config_bp.route('/cell', methods=['POST'])
@login_required
def update_cell():
    filename = request.json.get('file', 'config.xlsx')
    wb = get_workbook(filename)
    ws = wb.active
    row = int(request.json['row'])
    col = int(request.json['col'])
    value = request.json['value']
    ws.cell(row, col, value)
    save_workbook(wb, filename)
    wb.close()
    return jsonify({'success': True})

@config_bp.route('/cells', methods=['POST'])
@login_required
def update_cells():
    filename = request.json.get('file', 'config.xlsx')
    wb = get_workbook(filename)
    ws = wb.active
    changes = request.json['changes']
    for item in changes:
        row = int(item['row'])
        col = int(item['col'])
        value = item['value']
        ws.cell(row, col, value)
    save_workbook(wb, filename)
    wb.close()
    return jsonify({'success': True})

@config_bp.route('/row', methods=['POST'])
@login_required
def add_row():
    filename = request.json.get('file', 'config.xlsx')
    wb = get_workbook(filename)
    ws = wb.active
    count = int(request.json.get('count', 1))
    for _ in range(count):
        ws.append([''] * ws.max_column)
    save_workbook(wb, filename)
    wb.close()
    return jsonify({'success': True})

@config_bp.route('/row/<int:row_idx>', methods=['DELETE'])
@login_required
def delete_row(row_idx):
    filename = request.args.get('file', 'config.xlsx')
    wb = get_workbook(filename)
    ws = wb.active
    ws.delete_rows(row_idx)
    save_workbook(wb, filename)
    wb.close()
    return jsonify({'success': True})
